#!/usr/bin/env python3
#validator.py - Validates the data on S T R System/P E R System report for submission to LACOE

import openpyxl, pprint
from openpyxl.styles import Font, Color, PatternFill, Border
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.comments import Comment

from openpyxl.styles.colors import BLUE
from openpyxl.styles.colors import RED
from openpyxl.styles.colors import YELLOW

print('Opening workbook...')
wb = openpyxl.load_workbook('LA Co Of Ed.xlsx', data_only=True)
sheet = wb.active

end = 1
currentStart = 191001
currentEnd = 191031

#########################
#   DATA VALIDATION
#########################

#Find last row with data in column A
for lastRow in range(2, sheet.max_row):
    if not sheet['A' + str(lastRow)].value:
        end = lastRow
        break

#Color blank cells black & input text "BLANK"
for row in range(2, end):
    for col in range(1, column_index_from_string('AB')):
        if not sheet[get_column_letter(col) + str(row)].value:
            sheet[get_column_letter(col) + str(row)].fill = PatternFill(bgColor="FFC7CE", fill_type = "solid")
            sheet[get_column_letter(col) + str(row)].value = 'BLANK'

#Agency name (column A) should be the same value.  Uses value in A2 as correct #.
for row in range(3, end):
    if not (sheet['A' + str(2)].value == sheet['A' + str(row)].value):
        comment = Comment("Agency # does not match", "Windows User")
        sheet['A' + str(row)].fill = PatternFill(bgColor="FFC7CE", fill_type = "solid")
        sheet['A' + str(row)].comment = comment

#Gender must be either M or F
for row in range(2, end):    
    if not (sheet['D' + str(row)].value == 'M' or sheet['D' + str(row)].value == 'F'):
        comment = Comment("Gender must be M or F", "Windows User")
        sheet['D' + str(row)].fill = PatternFill(bgColor="FFC7CE", fill_type = "solid")
        sheet['D' + str(row)].comment = comment

#Classification must be 100010 or 200010
for row in range(2, end):    
    if not (sheet['H' + str(row)].value == 100010 or sheet['H' + str(row)].value == 200010):
        comment = Comment("Classification must be 100010 or 200010", "Windows User")
        sheet['H' + str(row)].fill = PatternFill(bgColor="FFC7CE", fill_type = "solid")
        sheet['H' + str(row)].comment = comment

#Transaction Code must be either TX, LX, RA or RX
for row in range(2, end):    
    if not (sheet['M' + str(row)].value == 'TX' or sheet['M' + str(row)].value == 'LX' or sheet['M' + str(row)].value == 'RX' or sheet['M' + str(row)].value == 'RA'):
        comment = Comment("Transaction code must be TX/LX/RA/RX", "Windows User")
        sheet['M' + str(row)].fill = PatternFill(bgColor="FFC7CE", fill_type = "solid")
        sheet['M' + str(row)].comment = comment

#Earning Type must be either REG, OVT, RTS
for row in range(2, end):    
    if not (sheet['N' + str(row)].value == 'REG' or sheet['N' + str(row)].value == 'OVT' or sheet['N' + str(row)].value == 'RTS'):
        comment = Comment("Earning type must be REG/OVT/RTS", "Windows User")
        sheet['N' + str(row)].fill = PatternFill(bgColor="FFC7CE", fill_type = "solid")
        sheet['N' + str(row)].comment = comment

#P E P R A Code must be either * or 1
for row in range(2, end):    
    if not (sheet['V' + str(row)].value == '*' or sheet['V' + str(row)].value == 1):
        comment = Comment("P E P R A Code must be */1", "Windows User")
        sheet['V' + str(row)].fill = PatternFill(bgColor="FFC7CE", fill_type = "solid")
        sheet['V' + str(row)].comment = comment

#Retirement Plan must be either S3, S5 or P9
for row in range(2, end):    
    if not (sheet['T' + str(row)].value == 'S3' or sheet['T' + str(row)].value == 'S5' or sheet['T' + str(row)].value == 'P9'):
        comment = Comment("Retirement Plan must be S3/S5/P9", "Windows User")
        sheet['T' + str(row)].fill = PatternFill(bgColor="FFC7CE", fill_type = "solid")
        sheet['T' + str(row)].comment = comment

#Retirement Status must be either M, N or R
for row in range(2, end):    
    if not (sheet['U' + str(row)].value == 'M' or sheet['U' + str(row)].value == 'N' or sheet['U' + str(row)].value == 'R'):
        comment = Comment("Retirement Status must be M/N/R", "Windows User")
        sheet['U' + str(row)].fill = PatternFill(bgColor="FFC7CE", fill_type = "solid")
        sheet['U' + str(row)].comment = comment

#Rate must be > 0
for row in range(2, end):    
    if not (sheet['R' + str(row)].value > 0):
        comment = Comment("Rate must be higher than zero", "Windows User")
        sheet['R' + str(row)].fill = PatternFill(bgColor="FFC7CE", fill_type = "solid")
        sheet['R' + str(row)].comment = comment

#Reporting Rate must be > 0
for row in range(2, end):    
    if not (sheet['AA' + str(row)].value > 0):
        comment = Comment("Reporting rate must be higher than zero", "Windows User")
        sheet['AA' + str(row)].fill = PatternFill(bgColor="FFC7CE", fill_type = "solid")
        sheet['AA' + str(row)].comment = comment

#Percentage must be > 0 && <= 100
for row in range(2, end):    
    if not (sheet['Z' + str(row)].value > 0 and sheet['Z' + str(row)].value <= 100):
        comment = Comment("Percentage higher than zero & less than or equal to 100", "Windows User")
        sheet['Z' + str(row)].fill = PatternFill(bgColor="FFC7CE", fill_type = "solid")
        sheet['Z' + str(row)].comment = comment

#Session must be S
for row in range(2, end):    
    if not (sheet['AB' + str(row)].value == 'S'):
        comment = Comment("Session must be S", "Windows User")
        sheet['AB' + str(row)].fill = PatternFill(bgColor="FFC7CE", fill_type = "solid")
        sheet['AB' + str(row)].comment = comment

#Number of pays must be 10, 11 or 12
for row in range(2, end):    
    if not (sheet['P' + str(row)].value == 10 or sheet['P' + str(row)].value == 11 or sheet['P' + str(row)].value == 12):
        comment = Comment("Number of pays must be 10, 11 or 12", "Windows User")
        sheet['P' + str(row)].fill = PatternFill(bgColor="FFC7CE", fill_type = "solid")
        sheet['P' + str(row)].comment = comment

#Time must be > 0
for row in range(2, end):    
    if not (sheet['Q' + str(row)].value > 0):
        comment = Comment("Time must be higher than zero", "Windows User")
        sheet['Q' + str(row)].fill = PatternFill(bgColor="FFC7CE", fill_type = "solid")
        sheet['Q' + str(row)].comment = comment

#########################
#        ALL
#########################

#Check current earnings
for row in range(2, end): 
    if (sheet['M' + str(row)].value == 'TX'):
        if not (sheet['I' + str(row)].value >= currentStart and sheet['I' + str(row)].value <= currentEnd):
            comment = Comment("Start date is not current", "Windows User")
            sheet['I' + str(row)].fill = PatternFill(fgColor=BLUE, fill_type = "solid")
            sheet['I' + str(row)].comment = comment

        if not(sheet['J' + str(row)].value >= currentStart and sheet['J' + str(row)].value <= currentEnd):
            comment = Comment("End date is not current", "Windows User")
            sheet['J' + str(row)].fill = PatternFill(fgColor=BLUE, fill_type = "solid")
            sheet['J' + str(row)].comment = comment

        if not (sheet['S' + str(row)].value > 0):
            comment = Comment("Earnings must be positive", "Windows User")
            sheet['S' + str(row)].fill = PatternFill(fgColor=BLUE, fill_type = "solid")
            sheet['S' + str(row)].comment = comment


        if not (sheet['AA' + str(row)].value > 0):
            comment = Comment("Reporting rate must be positive", "Windows User")
            sheet['AA' + str(row)].fill = PatternFill(fgColor=BLUE, fill_type = "solid")
            sheet['AA' + str(row)].comment = comment

#Check reversal earnings
for row in range(2, end): 
    if (sheet['M' + str(row)].value == 'RX'):
        if not (sheet['I' + str(row)].value < currentStart and sheet['I' + str(row)].value <= sheet['J' + str(row)].value):
            comment = Comment("Start date is not prior", "Windows User")
            sheet['I' + str(row)].fill = PatternFill(fgColor=BLUE, fill_type = "solid")
            sheet['I' + str(row)].comment = comment

        if not(sheet['J' + str(row)].value < currentStart and sheet['J' + str(row)].value >= sheet['I' + str(row)].value):
            comment = Comment("End date is not prior", "Windows User")
            sheet['J' + str(row)].fill = PatternFill(fgColor=BLUE, fill_type = "solid")
            sheet['J' + str(row)].comment = comment

        if not (sheet['S' + str(row)].value < 0):
            comment = Comment("Earnings must be negative (reversal)", "Windows User")
            sheet['S' + str(row)].fill = PatternFill(fgColor=BLUE, fill_type = "solid")
            sheet['S' + str(row)].comment = comment


        if not (sheet['AA' + str(row)].value > 0):
            comment = Comment("Reporting rate must be positive", "Windows User")
            sheet['AA' + str(row)].fill = PatternFill(fgColor=BLUE, fill_type = "solid")
            sheet['AA' + str(row)].comment = comment

#Check late earnings
for row in range(2, end): 
    if (sheet['M' + str(row)].value == 'LX'):
        if not (sheet['I' + str(row)].value < currentStart and sheet['I' + str(row)].value <= sheet['J' + str(row)].value):
            comment = Comment("Start date is not prior", "Windows User")
            sheet['I' + str(row)].fill = PatternFill(fgColor=BLUE, fill_type = "solid")
            sheet['I' + str(row)].comment = comment

        if not(sheet['J' + str(row)].value < currentStart and sheet['J' + str(row)].value >= sheet['I' + str(row)].value):
            comment = Comment("End date is not prior", "Windows User")
            sheet['J' + str(row)].fill = PatternFill(fgColor=BLUE, fill_type = "solid")
            sheet['J' + str(row)].comment = comment

        if not (sheet['S' + str(row)].value > 0):
            comment = Comment("Earnings must be positive", "Windows User")
            sheet['S' + str(row)].fill = PatternFill(fgColor=BLUE, fill_type = "solid")
            sheet['S' + str(row)].comment = comment


        if not (sheet['AA' + str(row)].value > 0):
            comment = Comment("Reporting rate must be positive", "Windows User")
            sheet['AA' + str(row)].fill = PatternFill(fgColor=BLUE, fill_type = "solid")
            sheet['AA' + str(row)].comment = comment

#Check stipends
for row in range(2, end): 
    if (sheet['O' + str(row)].value == 'L'):
        if not (sheet['R' + str(row)].value == sheet['S' + str(row)].value and sheet['S' + str(row)].value == sheet['AA' + str(row)].value):
            comment = Comment("This is a stipend.  Rate = earnings = reporting rate.", "Windows User")
            sheet['S' + str(row)].fill = PatternFill(fgColor=BLUE, fill_type = "solid")
            sheet['S' + str(row)].comment = comment


#Check monthly
socialSecurityNumList = []
for row in range(2, end):
    if (sheet['O' + str(row)].value == 'M'):
        if (sheet['M' + str(row)].value == 'TX'):
            #Check if it exists -- yes error, no add
            if sheet['F' + str(row)].value in socialSecurityNumList:
                comment = Comment("This is a duplicate monthly salary.  Verify this is correct, otherwise fix.  Should only have one salary line per month.", "Windows User")
                sheet['F' + str(row)].fill = PatternFill(fgColor=BLUE, fill_type = "solid")
                sheet['F' + str(row)].comment = comment
            else:
                socialSecurityNumList.append(sheet['F' + str(row)].value)
                

#########################
#      S T R SYSTEM
#########################

#Check hourly earnings
for row in range(2, end):
    
    #If a S T R System line
    if (sheet['H' + str(row)].value == 100010):
        
        #If an hourly line
        if (sheet['O' + str(row)].value == 'H'):
            
            if not (sheet['Q' + str(row)].value == (sheet['S' + str(row)].value/sheet['R' + str(row)].value)):
                comment = Comment("Time must be earnings/rate & positive", "Windows User")
                sheet['Q' + str(row)].fill = PatternFill(fgColor=RED, fill_type = "solid")
                sheet['Q' + str(row)].comment = comment
                
            if not (sheet['R' + str(row)].value < 100):
                comment = Comment("Rate must be an hourly rate, ideally less than $100/hr", "Windows User")
                sheet['R' + str(row)].fill = PatternFill(fgColor=RED, fill_type = "solid")
                sheet['R' + str(row)].comment = comment
                
            if not (sheet['AA' + str(row)].value > 25000):
                comment = Comment("Reporting rate must be positive & an annualized rate, ideally over 25k", "Windows User")
                sheet['AA' + str(row)].fill = PatternFill(fgColor=RED, fill_type = "solid")
                sheet['AA' + str(row)].comment = comment

#Check salary(monthly) earnings
for row in range(2, end):
    
    #If a S T R System line
    if (sheet['H' + str(row)].value == 100010):
        
        #If a monthly line
        if (sheet['O' + str(row)].value == 'M'):
                
            if not ((sheet['AA' + str(row)].value > 25000) or (sheet['R' + str(row)].value == sheet['AA' + str(row)].value)):
                comment = Comment("Reporting rate must be monthly or annualized", "Windows User")
                sheet['AA' + str(row)].fill = PatternFill(fgColor=RED, fill_type = "solid")
                sheet['AA' + str(row)].comment = comment

#Check retirement plans
for row in range(2, end):
    
    #If a S T R System line
    if (sheet['H' + str(row)].value == 100010):
        
        #If member earnings
        if (sheet['T' + str(row)].value == 'S5'):
            
            #Earnings should match subject
            if not (sheet['S' + str(row)].value == sheet['W' + str(row)].value):
                comment = Comment("Subject should equal earnings", "Windows User")
                sheet['W' + str(row)].fill = PatternFill(fgColor=RED, fill_type = "solid")
                sheet['W' + str(row)].comment = comment
                
                #If current or late
                if (sheet['U' + str(row)].value == 'M' and sheet['M' + str(row)].value == 'TX' or sheet['M' + str(row)].value == 'LX'):
                    if not (sheet['W' + str(row)].value > 0 and sheet['Y' + str(row)].value > 0 and sheet['Z' + str(row)].value > 0):
                        comment = Comment("Subject/EE/ER should be > zero", "Windows User")
                        sheet['W' + str(row)].fill = PatternFill(fgColor=RED, fill_type = "solid")
                        sheet['W' + str(row)].comment = comment
                        
                #If reversal
                if (sheet['U' + str(row)].value == 'M' and sheet['M' + str(row)].value == 'RX'):
                    if not (sheet['W' + str(row)].value < 0 and sheet['Y' + str(row)].value < 0 and sheet['Z' + str(row)].value < 0):
                        comment = Comment("Subject/EE/ER should be < zero", "Windows User")
                        sheet['W' + str(row)].fill = PatternFill(fgColor=RED, fill_type = "solid")
                        sheet['W' + str(row)].comment = comment

        #If nonmember earnings
        if (sheet['T' + str(row)].value == 'S3'):
            
            #Subject should be zero
            if not (sheet['W' + str(row)].value == 0):
                comment = Comment("Subject should be zero", "Windows User")
                sheet['W' + str(row)].fill = PatternFill(fgColor=RED, fill_type = "solid")
                sheet['W' + str(row)].comment = comment

                #If non-member or retiree
                if (sheet['U' + str(row)].value == 'N' or sheet['U' + str(row)].value == 'R'):
                    if not (sheet['W' + str(row)].value == 0 and sheet['Y' + str(row)].value == 0 and sheet['Z' + str(row)].value == 0):
                        comment = Comment("Subject/EE/ER should ALL be zero", "Windows User")
                        sheet['W' + str(row)].fill = PatternFill(fgColor=RED, fill_type = "solid")
                        sheet['W' + str(row)].comment = comment

#########################
#      P E R SYSTEM
#########################

#Check salary(monthly) earnings
for row in range(2, end):
    
    #If a P E R System line
    if (sheet['H' + str(row)].value == 200010):
        
        #If a monthly line
        if (sheet['O' + str(row)].value == 'M'):

            #Reporting rate should equal rate
            if not (sheet['R' + str(row)].value == sheet['AA' + str(row)].value):
                comment = Comment("Reporting rate should equal rate", "Windows User")
                sheet['AA' + str(row)].fill = PatternFill(fgColor=YELLOW, fill_type = "solid")
                sheet['AA' + str(row)].comment = comment

            #Subject should equal earned
            if not (sheet['W' + str(row)].value == sheet['S' + str(row)].value):
                comment = Comment("Subject should equal earned", "Windows User")
                sheet['W' + str(row)].fill = PatternFill(fgColor=YELLOW, fill_type = "solid")
                sheet['W' + str(row)].comment = comment

#Check hourly earnings
for row in range(2, end):
    
    #If a P E R System line
    if (sheet['H' + str(row)].value == 200010):
        
        #If an hourly line
        if (sheet['O' + str(row)].value == 'H'):

            #Reporting rate should equal rate
            if not (sheet['R' + str(row)].value == sheet['AA' + str(row)].value):
                comment = Comment("Reporting rate should equal rate", "Windows User")
                sheet['AA' + str(row)].fill = PatternFill(fgColor=YELLOW, fill_type = "solid")
                sheet['AA' + str(row)].comment = comment

            #Subject should equal earned
            if not (sheet['W' + str(row)].value == sheet['S' + str(row)].value):
                comment = Comment("Subject should equal earned", "Windows User")
                sheet['W' + str(row)].fill = PatternFill(fgColor=YELLOW, fill_type = "solid")
                sheet['W' + str(row)].comment = comment

            #Rate should be < $100/hr
            if not (sheet['R' + str(row)].value < 100):
                comment = Comment("Rate should be < $100/hr", "Windows User")
                sheet['R' + str(row)].fill = PatternFill(fgColor=YELLOW, fill_type = "solid")
                sheet['R' + str(row)].comment = comment

            #Time = earned/rate
            if not (sheet['Q' + str(row)].value == (sheet['S' + str(row)].value/sheet['R' + str(row)].value)):
                comment = Comment("Time = earned/rate", "Windows User")
                sheet['Q' + str(row)].fill = PatternFill(fgColor=YELLOW, fill_type = "solid")
                sheet['Q' + str(row)].comment = comment
                
wb.save('LA Co Of Ed - UPDATED.xlsx')
